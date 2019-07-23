import openpyxl
from openpyxl.styles import Font

wb = openpyxl.load_workbook('./data/example.xlsx')
sheet = wb.get_active_sheet()
cell = sheet.cell(1, 1)
print(cell.value)
print(cell.coordinate)

first_row = sheet[1]
last_row = sheet[sheet.max_row]

# Adding sum property at the end
new_row = sheet[sheet.max_row + 1]
new_row[0].value = 'Sum'
new_row[2].value = '=SUM(C1:C%s)' % last_row[0].row

sheet.merge_cells(':'.join(map(lambda x: x.coordinate, new_row[0:2])))

# Adding title at the top
new_row = sheet[sheet.max_row + 1]
new_row[0].value = 'Date'
new_row[1].value = 'Product'
new_row[2].value = 'Price'

for cell in new_row:
  cell.font = Font(b=True)

row_range = (
    ':'.join(map(lambda x: x.coordinate, [new_row[0], new_row[-1]])))
sheet.move_range(row_range, rows=-new_row[0].row + 1)

# Froze first row
sheet.freeze_panes = 'A2'

wb.save('./dist/example.xlsx')
