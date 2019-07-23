import openpyxl
from openpyxl.styles import Font
import sys

wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

N = int(sys.argv[1])

sheet.freeze_panes = 'B2'  # Row 1 and columns A and B

for row in range(1, N + 1):
  for column in range(1, N + 1):
    if row == 1 or column == 1:
      header_cells = None

      if row == 1 and column == 1:
        header_cells = [sheet.cell(row + 1, column),
                        sheet.cell(row, column + 1)]
      elif row == 1:
        header_cells = [sheet.cell(row, column + 1)]
      elif column == 1:
        header_cells = [sheet.cell(row + 1, column)]

      for header_cell in header_cells:
        header_cell.font = Font(b=True)
        header_cell.value = row * column

    sheet.cell(row + 1, column + 1).value = row * column

wb.save('./dist/mult-table.xlsx')
