import openpyxl
import os
import glob

for filename in glob.glob('**/*.xlsx'):
  full_filename = os.path.join(os.getcwd(), filename)

  wb = openpyxl.load_workbook(full_filename)
  sheet = wb.get_active_sheet()

  txt_filename = os.path.join(
      './dist', os.path.basename(filename).replace('.xlsx', '.txt'))

  for col in range(1, sheet.max_column):
    file = open(txt_filename, 'a')

    for row in range(1, sheet.max_row + 1):
      file.write(str(sheet.cell(row, col).value))

    file.close()
