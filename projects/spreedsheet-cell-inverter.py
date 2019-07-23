import openpyxl

current_wb = openpyxl.load_workbook('./data/items.xlsx')
current_sheet = current_wb.get_active_sheet()

new_wb = openpyxl.Workbook()
new_sheet = new_wb.get_active_sheet()

for row in range(1, current_sheet.max_row + 1):
  for col in range(1, current_sheet.max_column + 1):
    new_sheet.cell(col, row).value = current_sheet.cell(row, col).value

new_wb.save('./dist/items.xlsx')
