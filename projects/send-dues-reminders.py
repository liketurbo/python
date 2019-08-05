import openpyxl
import smtplib

wb = openpyxl.load_workbook('data/dues-records.xlsx')
sheet = wb.get_active_sheet()

lastest_month = sheet.cell(1, sheet.max_column).value

unpaid_members = []
for row_num in range(2, sheet.max_row + 1):
  payment = sheet.cell(row_num, sheet.max_column).value

  if payment != 'paid':
    name = sheet.cell(row_num, 1).value
    email = sheet.cell(row_num, 2).value

    unpaid_members.append((name, email))

smtp_obj = smtplib.SMTP('smtp.mail.ru', 587)
res = smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login('andrey.lukin.1986@bk.ru', 'Sswwf3i0NDyo4Pv')

for name, email in unpaid_members:
  body = 'Subject: {0} dues unpaid.\
\nDear {1},\nRecords show that you have not paid dues for {0}.\
\nPlease make this payment as soon as possible. Thank you!'.format(lastest_month, name)

  status = smtp_obj.sendmail('andrey.lukin.1986@bk.ru', email, body)

  if status != {}:
    print('There was problem sending email to %s: %s' % (email, status))
  else:
    print('Email sent to ' + email)
