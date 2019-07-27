import os
import openpyxl
import csv

os.makedirs('./dist/from-excel', exist_ok=True)

for excel_file in os.listdir('./data'):
  if not excel_file.endswith('.xlsx'):
    continue

  wb = openpyxl.load_workbook(os.path.join('./data', excel_file))

  csv_file = open(os.path.join('dist/from-excel',
                               os.path.splitext(excel_file)[0] + '.csv'), 'w')
  csv_writer = csv.writer(csv_file)

  for sheet_name in wb.sheetnames:
    sheet = wb.get_sheet_by_name(sheet_name)

    for row_num in range(1, sheet.max_row + 1):
      row_data = []

      for col_num in range(1, sheet.max_column + 1):
        row_data.append(sheet.cell(row_num, col_num).value)

      csv_writer.writerow(row_data)

  csv_file.close()
